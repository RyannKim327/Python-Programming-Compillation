import openpyxl as xl
from openpyxl.workbook import Workbook
import os

# to make it constant filename and sheet name
filename = "book.xlsx"
sheet_name = "book"

# file check if existed or not
if os.path.exists(filename):
	# if existed, just load it
	wb = xl.load_workbook(filename)
	ws = wb[sheet_name]
else:
	# if not, create new workbook
	wb = Workbook()
	# this is just to create a new sheet if there's none
	ws = wb.active
	# modify the name, never forget this
	ws.title = sheet_name

	# create some columns here [optional]
	column = (
		"ID",
		"Name",
		"Publish Date",
		"Author"
	)

	ws.append(column)

# add some data
data = (
	(
		"BOOK_001",
		"Harry Potter and the Sorcerer's Stone",
		"19997",
		"J.K Rowling"
	),
	(
		"BOOK_002",
		"Harry Potter and the Goblet of Fire",
		"2000",
		"J.K Rowling"
	)
)

# you may create multiple tuples here, it's up to you

for i in data:
	ws.append(i)

# this is just optional
limit = 2

for i in range(len(list(ws.iter_rows(values_only=True))) - 1, -1, -1):
	for j in list(ws.iter_rows(values_only=True))[i]:
		print(j, end="\t")
	print()
	# to have some limits to print
	limit -= 1
	if limit == 0:
		break

# save as xlsx file
wb.save(filename)