from openpyxl import *
import os

def createDatabase(data: list):
	global xls, wb

	if os.path.exists("book.xlsx"):
		wb = load_workbook("book.xlsx")
		xls = wb['currentdatabase']
	else:
		wb = Workbook()
		xls = wb.active
		xls.title = "currentdatabase"
		xls.cell(row = 1, column = 1, value = "ID")
		for i in range(0, len(data)):
			xls.cell(row = 1, column = i + 2, value = data[i])

def addData(data: list):
	rows = xls.max_row + 1
	xls.cell(row = rows, column = 1, value = rows - 1)
	for i in range(0, len(data)):
		xls.cell(row = rows, column = i + 2, value = data[i])
	
	wb.save("book.xlsx")

def getData():
	data = []
	for r in xls.iter_rows(values_only=True):
		d = []
		for c in r:
			d.append(c)
		data.append(d)

	return data