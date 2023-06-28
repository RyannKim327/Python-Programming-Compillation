from openpyxl import *
import os

def createDatabase():
	global wb, ws, fn

	fn = "employees.xlsx"

	if os.path.exists(fn):
		wb = load_workbook(fn)
		ws = wb['employees']
	else:
		wb = Workbook()
		ws = wb.active
		ws.title = "employees"
		columns = [
			"reference",
			"name",
			"email",
			"gender",
			"destination",
			"contact",
			"salary",
			"address"
		]
		for i in range(len(columns)):
			ws.cell(row = 1, column = 1 + i, value = columns[i])


def addData(data: dict):
	proceed = True
	invalids = []
	send = []
	if data.get("reference") == None or data['reference'] == "":
		invalids.append("reference")
		proceed = False
	else:
		if data['reference'] in getAllID():
			return {
				"approve": False,
				"invalids": f"The Reference ID {data['reference']} is already existed."
			}
		else:
			send.append(data['reference'])

	if data.get("name") == None or data['name'] == "":
		invalids.append("name")
		proceed = False
	else:
		send.append(data['name'])
	if data.get("email") == None or data['email'] == "":
		invalids.append("email")
		proceed = False
	else:
		send.append(data['email'])

	if data.get("gender") == None or data['gender'] == "":
		invalids.append("gender")
		proceed = False
	else:
		send.append(data['gender'])

	if data.get("destination") == None or data['destination'] == "":
		invalids.append("destination")
		proceed = False
	else:
		send.append(data['destination'])

	if data.get("contact") == None or data['contact'] == "" or not data['contact'].isdigit():
		invalids.append("contact")
		proceed = False
	else:
		send.append(data['contact'])

	if data.get("salary") == None or data['salary'] == "" or not data['salary'].isdigit():
		invalids.append("salary")
		proceed = False
	else:
		send.append(data['salary'])

	if data.get("address") == None or data['address'] == "":
		invalids.append("address")
		proceed = False
	else:
		send.append(data['address'])

	if proceed:
		rows = ws.max_row + 1
		for i in range(len(send)):
			ws.cell(row = rows, column = i + 1, value = send[i])
		wb.save(fn)

	
	return {
		"approve": proceed,
		"invalids": ", ".join(invalids)
	}

def getData(search = ""):
	data = []
	if search == "":
		x = False
		for r in ws.iter_rows(values_only=True):
			if x:
				d = []
				for i in r:
					d.append(i)
				if d[0] != None:
					data.append(d)
			x = True
	else:
		y = False
		for r in ws.iter_rows(values_only=True):
			if y:
				d = []
				x = False
				for i in r:
					if search.lower() in i.lower():
						x = True
						break
				if x:
					for i in r:
						d.append(i)
					if d[0] != None:
						data.append(d)
			y = True
	# for c in ws.iter_cols(values_only=True):
	# 	d = []
	# 	for r in c:
	# 		d.append(r)
	# 	data.append(d)
	return data

def deleteData(ids: list):
	x = 1
	for j in ws.iter_rows(values_only=True):
		if j[0] in ids:
			ws.delete_rows(x)
		x += 1
	ws.save(fn)


def getAllID():
	data = []
	for r in ws.iter_rows(values_only=True):
		data.append(r[0])
	return data

def lastID():
	return f"EMP_{len(list(ws.iter_rows(values_only=True)))}"

def updateData(id: str, data: list):
	employee_id = 0
	x = 1
	for i in ws.iter_rows(values_only=True):
		if i[0] == id:
			employee_id = x
			break
		x += 1

	for i in range(len(data)):
		if data[i] != "":
			ws.cell(row = employee_id, column = i + 2, value = data[i])
	
	wb.save(fn)