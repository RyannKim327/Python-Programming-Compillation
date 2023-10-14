from openpyxl.workbook import Workbook
from openpyxl import load_workbook

import os


filename = "sample.xlsx"

if os.path.exists(filename):
	excel_con = load_workbook(filename)
	excel_activate = excel_con['sample']
	print("existed")
else:
	excel_con = Workbook()
	excel_activate = excel_con.active # Sheet
	excel_activate.title = "sample"
	excel_activate.append((
		"ID",
		"Name",
		"Age"
	)
	)
	print("not existed")

excel_con.save("sample.xlsx")

for i in excel_activate.iter_rows(values_only=True):
	print(i)