from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import os, hashlib

# Encryption
def encrypt(password: str) -> str:
	return (hashlib.sha256(password.encode())).hexdigest()

# Save
def save():
	wb.save(filename)

# Questions
def createExcell():
	global wb, questions, filename, archieves
	filename = "data.xlsx"
	quest = "questions"
	archs = "archieves questions"
	if os.path.exists(filename):
		wb = load_workbook(filename)
	else:
		wb = Workbook()
	
	if quest in wb.sheetnames:
		questions = wb[quest]
	else:
		questions = wb.active
		questions.title = quest

		columns = (
			"question",
			"answer",
			"case sensitive",
			"question by"
		)
		questions.append(columns)
	
	if archs in wb.sheetnames:
		archieves = wb[archs]
	else:
		archieves = wb.create_sheet(archs)
		columns = (
			"question",
			"answer",
			"case sensitive",
			"question by"
		)
		archieves.append(columns)		
	
	save()

def addQuestion(q: str, a: str, cs: bool, t: str):
	questions.append((q, a, cs, t))
	save()
	return True

def getAllQuestions():
	lists = []
	for i in questions.iter_rows(values_only=True):
		lists.append(i)
	return lists

def updateQuestion(pos: int, data: list):
	for i in range(len(data)):
		questions.cell(row=pos, column=i + 1, value=data[i])
	save()
	return {
		"done": True,
		"msg": "Data modified successfully"
	}

def archieveQuestions(pos: int):
	x = 0

	for i in questions.iter_rows(values_only=True):
		if x == pos:
			archieves.append(i)
			break
		x += 1

	questions.delete_rows(pos + 1)
	save()
	return {
		"done": True,
		"msg": "Data archieved successfully"
	}

def listsArchieve():
	lists = []
	x = False
	for i in archieves.iter_rows(values_only=True):
		if x:
			lists.append(i)
		x = True
	
	return lists

def retrieveQuestion(pos: int):
	x = 0

	for i in archieves.iter_rows(values_only=True):
		if x == pos:
			questions.append(i)
			break
		x += 1

	archieves.delete_rows(pos + 1)
	save()
	return {
		"done": True,
		"msg": "Data retrieved successfully"
	}

# Students
def createStudents():
	global students
	studs = "students"
	
	if studs in wb.sheetnames:
		students = wb[studs]
	else:
		students = wb.create_sheet(studs)

		columns = (
			"ID",
			"fullname",
			"password",
			"score"
		)
		students.append(columns)
	
	save()

def addStudent(randomID: str, fullname: str, password: str) -> dict:
	id = randomID
	n = fullname
	p = encrypt(password)
	exist = False

	for i in students.iter_rows(values_only=True):
		if id in i:
			exist = True
			break
	
	if not exist:
		students.append((
			id,
			n,
			p,
			0
		))
		save()
	
	return {
		"exists": exist
	}

def getStudentId(username: str, password: str) -> dict:
	pos = 0
	userID = ""
	done = False
	for i in students.iter_rows(values_only=True):
		if i[0] == username and i[2] == encrypt(password):
			done = True
			userID = i[0]
			break
		pos += 1
	return {
		"done": done,
		"userID": userID,
		"pos": pos
	}


def updateScore(id: str, score: int):
	x = 0
	for i in students.iter_rows(values_only=True):
		if id == i[0]:
			break
		x += 1
	students.cell(row=x + 1, column=4, value=score)
	save()


def leaderboard():
	lists = []
	done = False
	x = False
	for i in students.iter_rows(values_only=True):
		if x:
			lists.append(list(i))
		x = True
	
	for i in range(len(lists)):
		for j in range(i):
			if int(lists[i][3]) > int(lists[j][3]):
				lists[i], lists[j] = lists[j], lists[i]
		
	return {
		"done": done,
		"lists": lists
	}


def removeStudent(pos: int):
	students.delete_rows(pos + 1)
	save()
	return {
		"done": True,
		"mgs": "A student was deleted successfully"
	}


# Teachers
def createTeachers():
	global teachers
	teach = "teachers"
	
	if teach in wb.sheetnames:
		teachers = wb[teach]
	else:
		teachers = wb.create_sheet(teach)

		columns = (
			"ID",
			"fullname",
			"password"
		)
		teachers.append(columns)
	
	save()

def addTeacher(randomID: str, fullname: str, password: str) -> dict:
	id = randomID
	n = fullname
	p = encrypt(password)
	exist = False

	for i in teachers.iter_rows(values_only=True):
		if id in i:
			exist = True
			break
	
	if not exist:
		teachers.append((
			id,
			n,
			p
		))
		save()
	
	return {
		"exists": exist
	}

def getTeacherId(username: str, password: str) -> dict:
	userID = 1
	done = False
	for i in teachers.iter_rows(values_only=True):
		if i[0] == username and i[2] == encrypt(password):
			done = True
			break
		userID += 1
	return {
		"done": done,
		"userID": userID
	}

def getTeacher(id: int):
	teacher = {}
	for i in range(len(teachers[id])):
		teacher[teachers[1][i].value] = teachers[id][i].value
	return teacher

def removeTeacher(pos: int):
	teachers.delete_rows(pos + 1)
	save()
	return {
		"done": True,
		"mgs": "A teacher was deleted successfully"
	}

# All users
def getAllUsers(_type, sorted: bool = True):
	lists = []
	if _type == 'teacher':
		x = False
		for i in teachers.iter_rows(values_only = True):
			l = []
			if x:
				for j in range(0, 2):
					l.append(i[j])
				lists.append(l)
			x = True
	else:
		x = False
		for i in students.iter_rows(values_only = True):
			l = []
			if x:
				for j in range(len(i)):
					if j != 2:
						l.append(i[j])
				lists.append(l)
			x = True
		if sorted:
			for l in range(len(lists)):
				for m in range(l):
					if lists[l][2] > lists[m][2]:
						lists[l], lists[m] = lists[m], lists[l]
	return lists
