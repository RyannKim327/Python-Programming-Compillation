import os
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from tkinter import messagebox

def toPDF(excelFile: str, sheetname: str = "",  pdfFile: str = ""):
	result = {
		"done": True,
		"msg": []
	}
	error = []
	if not excelFile.endswith(".xlsx"):
		result['done'] =  False
		error.append("excell file name")
	if not pdfFile.endswith(".pdf"):
		result['done'] = False
		error.append("pdf file name")
	
	result['mgs'] = error
	if result['done']:
		excel_path = os.path.abspath(excelFile)
		pdf_path = os.path.abspath(pdfFile)

		book = load_workbook(excel_path)
		if sheetname == "":
			sheet = book.active
		else:
			sheet = book[sheetname]
		
		mr_ = sheet.max_row
		mc = sheet.max_column
		
		can = canvas.Canvas(pdf_path, pagesize=landscape(letter))
		x = mr_
		max_data = 25
		x2 = []
		while x > 0:
			x2.append(x - (x - max_data))
			x -= max_data

		x = 1
		r = 1

		for i in range(len(x2)):
			can.drawString(1, 5, f"{sheetname.capitalize()}: ({i + 1})")

			mr = int(x + x2[i])
			x += mr

			tm = 3 * inch
			lm = 0.75 * inch
			bm = 0.75 * inch
			rm = 0.75 * inch

			cell_width = (11 * inch - lm - rm) / mc 
			cell_height = (8.5 * inch - tm - bm) / (mr_ // len(x2))
			
			for row in range(r, r + max_data):
				if max_data > mr_ + 1:
					break
				for col in range(1, mc + 1):
					cell  = sheet.cell(row=row, column=col)
					txt = str(cell.value)

					if col == 4:
						col += 1
					if "question" in sheetname:
						if col >= 2:
							x = lm + (col + 1) * cell_width
							# y = 11 * inch - (tm + ((row % 10) + 1) * cell_height)
						else:
							x = lm + (col - 1) * cell_width
							# y = 11 * inch - (tm + ((row % 10) + 1) * cell_height)
					else:
						x = lm + (col - 1) * cell_width
					
					y = 11 * inch - (tm + (row - r)  * cell_height)

					if col != 3:
						if txt != "None":
							can.drawString(x, y, txt)
					
			r += max_data

			can.showPage()	

		can.save()

		result['msg'] = "Success"
	return result