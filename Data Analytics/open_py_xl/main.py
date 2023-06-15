from openpyxl import *

class xl:
	def __init__(self, filename: str = ""):
		"""
			Project initialization, this is to set some files here
		"""
		if filename == "":
			self.filename = "Book.xlsx"
		else:
			self.filename = filename
		try:
			self.xl = load_workbook(self.filename)
		except:
			self.xl = Workbook()

	def create_s(self, sheet_name: str = ""):
		"""
			Create a sheet, where ID is automatically generated
		"""
		self.xl_s = self.xl.active
		if not sheet_name == "":
			self.xl_s.title = sheet_name
		
		self.xl_s.cell(row = 1, column = 1, value = "ID")
	
	def set_column(self, cols: list):
		"""
			Set some columns or column names
		"""
		for a in range(len(cols)):
			self.xl_s.cell(row = 1, column = a + 2, value = cols[a])
	
	def read_data(self, position):
		"""
			Reading data
		"""
		return self.xl_s[position]
	
	def addData(self, data: list = [[]]):
		"""
			Add a data thru 2d array like:
			[
				[],
				[],
				[]
			]
		"""
		for a in range(len(data)):
			self.xl_s.cell(row = a + 2, column = 1, value = a + 1)
			for b in range(len(data[a])):
				self.xl_s.cell(row = a + 2, column = b + 2, value = data[a][b])

	def save_it(self, filename: str = ""):
		"""
			Save as xlsx
		"""
		if filename == "" or not "." in filename:
			self.xl.save(self.filename)
		else:
			self.xl.save(filename)

# Start Here
if __name__ == "__main__":
	data = xl()
	data.create_s("Sample")
	data.set_column([
		"Name",
		"Description",
		"Gender"
	])
	data.addData([
		[
			"Ryann Kim Sesgundo",
			"A newbie programmer from Quezon Province",
			"Male"
		],
		[
			"Ryann Kim Sesgundo",
			"A newbie programmer from Quezon Province",
			"Male"
		],
	])
	data.save_it("Book.xlsx")